"""
make_palette_xlsx.py — regenerate ArcGIS_Pro_Color_Palette.xlsx from palette.csv

This script rebuilds the workbook with three sheets:
- Palette Grid
- Color Reference
- About

Default behavior reads from and writes to:
  ../arcgis-pro-standard-palette
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import tempfile
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_DIR = SCRIPT_DIR.parent / "arcgis-pro-standard-palette"

PALETTE_CSV_NAME = "palette.csv"
OUTPUT_XLSX_NAME = "ArcGIS_Pro_Color_Palette.xlsx"

ABOUT_LINES = [
    "About this palette reference",
    "",
    "Sources:",
    "  Esri Knowledge Base 000010027 — 'What Are the RGB Color Values for the Standard ArcMap Color Set?'",
    "    https://support.esri.com/en-us/knowledge-base/what-are-the-rgb-color-values-for-the-standard-arcmap-c-000010027",
    "  Esri ArcGIS Colors (ArcGIS Pro 3.6, 2025) — 'ArcGIS_Colors.pdf' from the ArcGIS Colors system style",
    "    https://pro.arcgis.com/en/pro-app/latest/help/mapping/layer-properties/color.htm",
    "",
    "Names and RGB values match Esri's current ArcGIS Pro 3.6 ArcGIS Colors reference (2025).",
    "Esri corrected several spellings between the original ArcMap palette and the current ArcGIS Pro palette:",
    "  Fushia -> Fuchsia     (in 'Fuchsia Pink' and 'Medium Fuchsia')",
    "  Cretean -> Cretan     (in 'Cretan Blue')",
    "  Chrysophase -> Chrysoprase",
    "  Citroen -> Citron     (in 'Citron Yellow')",
    "",
    "The palette contains 120 named colors arranged in a 12-column, 10-row grid.",
    "Column 1 is the grayscale ramp (Arctic White through Black); each row cycles through",
    "the same 11 hues at progressively darker, more saturated, or more muted values.",
    "",
    "Structure by row:",
    "  Row 1:  Very pale pastels (near-white wash of each hue)",
    "  Row 2:  Gray 10% + light brights",
    "  Row 3:  Gray 20% + pure saturated primaries",
    "  Row 4:  Gray 30% + deep saturated",
    "  Row 5:  Gray 40% + earthy mediums",
    "  Row 6:  Gray 50% + deep earth tones",
    "  Row 7:  Gray 60% + dusty pastels (muted lights)",
    "  Row 8:  Gray 70% + medium muted",
    "  Row 9:  Gray 80% + deeper muted",
    "  Row 10: Black + darkest deep tones",
    "",
    "Descriptions (5-10 words each) combine three angles:",
    "  - Functional: lightness (pale/medium/deep/very dark), warmth (warm/cool/balanced), saturation (pure/muted/earthy)",
    "  - Evocative: a familiar real-world anchor (mango flesh, old denim, glacier ice)",
    "  - Cartographic: a typical mapping use (water, vegetation, residential fills, boundaries)",
    "",
    "These three properties — lightness, warmth, and saturation — are more reliably",
    "perceivable across different types of color vision than hue alone.",
]


def load_palette(csv_path: Path) -> list[dict]:
    if not csv_path.exists():
        raise FileNotFoundError(f"palette CSV not found: {csv_path}")

    rows: list[dict] = []
    with csv_path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        required = {
            "index",
            "name",
            "r",
            "g",
            "b",
            "hex",
            "h_deg",
            "s_pct",
            "v_pct",
            "grid_row",
            "grid_col",
            "description",
        }
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"palette.csv missing columns: {sorted(missing)}")

        for row in reader:
            rows.append(
                {
                    "index": int(row["index"]),
                    "name": row["name"],
                    "r": int(row["r"]),
                    "g": int(row["g"]),
                    "b": int(row["b"]),
                    "hex": row["hex"].lstrip("#").upper(),
                    "h_deg": float(row["h_deg"]),
                    "s_pct": float(row["s_pct"]),
                    "v_pct": float(row["v_pct"]),
                    "grid_row": int(row["grid_row"]),
                    "grid_col": int(row["grid_col"]),
                    "description": row["description"],
                }
            )

    rows.sort(key=lambda r: r["index"])
    if len(rows) != 120:
        raise ValueError(f"Expected 120 colors, found {len(rows)}")
    return rows


def _text_color_for_bg(r: int, g: int, b: int) -> str:
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255.0
    return "000000" if luminance > 0.55 else "FFFFFF"


def build_palette_grid_sheet(wb: Workbook, palette: list[dict]) -> None:
    ws = wb.create_sheet("Palette Grid")

    ws.merge_cells("A1:L1")
    ws.merge_cells("A2:L2")
    ws["A1"] = "Esri (ArcMap/ArcGIS Pro) Standard Color Palette (120 colors)"
    ws["A2"] = "Names per ArcGIS Pro 3.6 (2025)  •  RGB values per Esri KB 000010027  •  See 'Color Reference' sheet for a sortable table"

    ws.freeze_panes = "A4"
    ws.row_dimensions[1].height = 24

    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    subtitle_fill = PatternFill(fill_type="solid", fgColor="2E75B6")

    for col in range(1, 13):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 19
        ws.cell(1, col).fill = title_fill
        ws.cell(2, col).fill = subtitle_fill

    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A2"].font = Font(color="FFFFFF", bold=False, size=10)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )

    for color in palette:
        base_row = 4 + (color["grid_row"] - 1) * 4
        col = color["grid_col"]

        name_row = base_row
        rgb_row = base_row + 1
        hex_row = base_row + 2
        desc_row = base_row + 3

        ws.row_dimensions[name_row].height = 20
        ws.row_dimensions[rgb_row].height = 14
        ws.row_dimensions[hex_row].height = 14
        ws.row_dimensions[desc_row].height = 36

        fill = PatternFill(fill_type="solid", fgColor=color["hex"])
        text_color = _text_color_for_bg(color["r"], color["g"], color["b"])

        ws.cell(name_row, col, color["name"])
        ws.cell(rgb_row, col, f"{color['r']}, {color['g']}, {color['b']}")
        ws.cell(hex_row, col, f"#{color['hex']}")
        ws.cell(desc_row, col, color["description"])

        for r in (name_row, rgb_row, hex_row, desc_row):
            cell = ws.cell(r, col)
            cell.fill = fill
            cell.font = Font(color=text_color, bold=(r == name_row), size=9 if r == name_row else 8)
            if r == desc_row:
                cell.font = Font(color=text_color, size=8, italic=True)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_color_reference_sheet(wb: Workbook, palette: list[dict]) -> None:
    ws = wb.create_sheet("Color Reference")
    headers = ["#", "Row", "Col", "Name", "R", "G", "B", "Hex", "H°", "S%", "V%", "Swatch", "Description"]
    ws.append(headers)

    widths = {
        "A": 5,
        "B": 6,
        "C": 6,
        "D": 22,
        "E": 6,
        "F": 6,
        "G": 6,
        "H": 10,
        "I": 8,
        "J": 8,
        "K": 8,
        "L": 12,
        "M": 52,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:M121"

    border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )

    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    for col in range(1, 14):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, color in enumerate(palette, start=2):
        ws.row_dimensions[idx].height = 22
        ws.cell(idx, 1, color["index"])
        ws.cell(idx, 2, color["grid_row"])
        ws.cell(idx, 3, color["grid_col"])
        ws.cell(idx, 4, color["name"])
        ws.cell(idx, 5, color["r"])
        ws.cell(idx, 6, color["g"])
        ws.cell(idx, 7, color["b"])
        ws.cell(idx, 8, f"#{color['hex']}")
        ws.cell(idx, 9, color["h_deg"])
        ws.cell(idx, 10, color["s_pct"])
        ws.cell(idx, 11, color["v_pct"])
        ws.cell(idx, 12, "")
        ws.cell(idx, 13, color["description"])

        for col in range(1, 14):
            cell = ws.cell(idx, col)
            cell.border = border
            if col == 4 or col == 13:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(idx, 12).fill = PatternFill(fill_type="solid", fgColor=color["hex"])


def build_about_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("About")
    ws.column_dimensions["A"].width = 110

    for row_idx, line in enumerate(ABOUT_LINES, start=1):
        ws.cell(row_idx, 1, line)
        cell = ws.cell(row_idx, 1)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if row_idx == 1:
            cell.font = Font(bold=True, size=16)
        elif line == "Sources:" or line == "Structure by row:" or line.startswith("Descriptions"):
            cell.font = Font(bold=True)
        else:
            cell.font = Font(size=11)


def generate_workbook(palette_csv: Path, output_xlsx: Path) -> None:
    palette = load_palette(palette_csv)
    wb = Workbook()
    active_sheet = wb.active
    if active_sheet is not None:
        wb.remove(active_sheet)

    build_palette_grid_sheet(wb, palette)
    build_color_reference_sheet(wb, palette)
    build_about_sheet(wb)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def _xlsx_content_digest(xlsx_path: Path) -> str:
    """Return a content digest for an xlsx, excluding volatile core metadata."""
    hasher = hashlib.sha256()
    with ZipFile(xlsx_path, "r") as zf:
        names = sorted(name for name in zf.namelist() if name != "docProps/core.xml")
        for name in names:
            hasher.update(name.encode("utf-8"))
            hasher.update(b"\0")
            hasher.update(zf.read(name))
            hasher.update(b"\0")
    return hasher.hexdigest()


def check_workbook_up_to_date(palette_csv: Path, output_xlsx: Path) -> bool:
    if not output_xlsx.exists():
        return False

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir) / OUTPUT_XLSX_NAME
        generate_workbook(palette_csv, tmp_path)
        return _xlsx_content_digest(output_xlsx) == _xlsx_content_digest(tmp_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Regenerate ArcGIS_Pro_Color_Palette.xlsx from palette.csv")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_DIR,
        help=f"Folder containing {PALETTE_CSV_NAME} (default: {DEFAULT_DIR})",
    )
    parser.add_argument(
        "--output-file",
        type=Path,
        default=DEFAULT_DIR / OUTPUT_XLSX_NAME,
        help=f"Output workbook path (default: {DEFAULT_DIR / OUTPUT_XLSX_NAME})",
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="Check whether output workbook is up to date; exits 1 if regeneration is needed",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    palette_csv = args.input_dir / PALETTE_CSV_NAME
    output_xlsx = args.output_file

    try:
        if args.check:
            if check_workbook_up_to_date(palette_csv, output_xlsx):
                print(f"Up to date: {output_xlsx}")
                return 0
            print(f"Out of date: {output_xlsx}")
            print("Run without --check to regenerate the workbook.")
            return 1

        generate_workbook(palette_csv, output_xlsx)
    except (FileNotFoundError, ValueError) as exc:
        print(f"ERROR: {exc}")
        return 1

    print(f"Wrote workbook: {output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
